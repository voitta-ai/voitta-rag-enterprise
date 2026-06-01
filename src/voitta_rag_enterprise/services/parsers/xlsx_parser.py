"""Excel parser — modern ``.xlsx``/``.xlsm`` via openpyxl, legacy
``.xls`` via xlrd.

Both paths emit the same shape: one markdown section per worksheet
keyed by ``## Sheet: <name>`` (the natural chunk boundary the chunker
already prefers). Row/column caps come from indexing_caps so a
spreadsheet with hundreds of thousands of rows doesn't drown the chunker.

Per-chunk sheet attribution would require threading metadata through
the chunker; out of scope for v1. For files synced from Google Drive
the user gets sheet names in the markdown headings, and the file's
Drive URL in the chunk payload.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import ClassVar

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..indexing_caps import get_caps
from ._ooxml import MIN_IMG_DIM, image_dimensions, iter_package_media
from .base import BaseParser, ExtractedImage, ParserResult

logger = logging.getLogger(__name__)


def _harvest_sheet_images(file_path: Path) -> list[ExtractedImage]:
    """Embedded pictures from an ``.xlsx``/``.xlsm`` package.

    openpyxl's ``read_only`` mode (used for the text pass to stay memory-safe on
    huge sheets) doesn't load drawings, so images are read straight from the OPC
    package instead. They float over cells with no stable text offset, so all
    land at position 0. Tiny decorative glyphs are filtered out.
    """
    images: list[ExtractedImage] = []
    for _name, blob, mime in iter_package_media(file_path, media_prefix="xl/media/"):
        width, height = image_dimensions(blob)
        if width and height and max(width, height) < MIN_IMG_DIM:
            continue
        images.append(
            ExtractedImage(bytes=blob, mime=mime, position=0, width=width, height=height)
        )
    return images


class XlsxParser(BaseParser):
    extensions: ClassVar[list[str]] = [".xlsx", ".xlsm", ".xls"]

    def parse(self, file_path: Path) -> ParserResult:
        if file_path.suffix.lower() == ".xls":
            return _parse_xls(file_path)
        try:
            wb = load_workbook(
                str(file_path),
                read_only=True,
                data_only=True,  # use cached formula results, not "=A1+B1"
            )
        except Exception as e:
            return ParserResult.failure(f"xlsx open failed: {e}")

        caps = get_caps()
        max_rows = caps.xlsx_max_rows
        max_cols = caps.xlsx_max_cols
        sections: list[str] = []
        try:
            for sheet in wb.worksheets:
                if sheet.sheet_state != "visible":
                    continue
                rendered = _render_sheet(sheet, max_rows=max_rows, max_cols=max_cols)
                if rendered.strip():
                    sections.append(f"## Sheet: {sheet.title}\n\n{rendered}")
        finally:
            wb.close()

        # Embedded pictures (logos, screenshots, photos placed on sheets). The
        # text pass above runs read-only and never sees them, so harvest from
        # the package. Charts are vector/XML, not rasters, so they're not
        # included. Legacy ``.xls`` (handled in ``_parse_xls``) is BIFF, not a
        # zip, so it has no media-harvest path.
        images = _harvest_sheet_images(file_path)
        return ParserResult(content="\n\n".join(sections), images=images)


def _parse_xls(file_path: Path) -> ParserResult:
    """Legacy ``.xls`` (BIFF) workbooks via xlrd.

    Same row/column caps as the modern path, same output shape. xlrd
    2.x dropped .xlsx support — it's strictly for .xls here.
    """
    try:
        import xlrd  # noqa: PLC0415 — optional dep, lazy import
    except ImportError:
        return ParserResult.failure(
            "xlrd not installed (needed for .xls). pip install xlrd."
        )
    try:
        wb = xlrd.open_workbook(str(file_path))
    except Exception as e:
        return ParserResult.failure(f"xls open failed: {e}")

    caps = get_caps()
    max_rows = caps.xlsx_max_rows
    max_cols = caps.xlsx_max_cols
    sections: list[str] = []
    for sheet_name in wb.sheet_names():
        sheet = wb.sheet_by_name(sheet_name)
        rows: list[list[str]] = []
        for row_idx in range(min(sheet.nrows, max_rows)):
            row = [
                _cell_to_str(sheet.cell_value(row_idx, col_idx))
                for col_idx in range(min(sheet.ncols, max_cols))
            ]
            rows.append(row)
        rendered = _format_table(rows)
        if rendered.strip():
            sections.append(f"## Sheet: {sheet_name}\n\n{rendered}")
    return ParserResult(content="\n\n".join(sections))


def _render_sheet(sheet: Worksheet, *, max_rows: int, max_cols: int) -> str:
    """Render a worksheet as a pipe-style markdown table.

    Trims trailing blank rows/columns so a sheet with five real rows but
    1,000 empty placeholders doesn't drown the chunker. The first row is
    treated as a header — that's the convention almost every spreadsheet
    follows and gives the downstream model a free signal about column
    semantics.
    """
    rows: list[list[str]] = []
    for row_index, raw in enumerate(sheet.iter_rows(values_only=True)):
        if row_index >= max_rows:
            break
        row = [_cell_to_str(c) for c in raw[:max_cols]]
        rows.append(row)
    return _format_table(rows)


def _format_table(rows: list[list[str]]) -> str:
    """Render a 2D ``str`` matrix as a pipe-style markdown table.

    Trims trailing blank rows/columns so a sheet with five real rows but
    1,000 empty placeholders doesn't drown the chunker. The first row is
    treated as a header — that's the convention almost every spreadsheet
    follows. Pure helper so both the openpyxl and xlrd paths share it.
    """
    rows = _trim_trailing_blank(rows)
    if not rows:
        return ""

    width = max(len(r) for r in rows)
    for r in rows:
        while len(r) < width:
            r.append("")

    while width > 1 and all(not r[width - 1] for r in rows):
        width -= 1
        for r in rows:
            r.pop()
    if width == 0:
        return ""

    header = rows[0]
    body = rows[1:]
    lines = ["| " + " | ".join(_md_escape(c) for c in header) + " |"]
    lines.append("| " + " | ".join(["---"] * width) + " |")
    for r in body:
        lines.append("| " + " | ".join(_md_escape(c) for c in r) + " |")
    return "\n".join(lines)


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, str):
        return value.strip()
    return str(value)


def _md_escape(text: str) -> str:
    """Pipe-tables are line-based; escape pipes and collapse newlines."""
    return text.replace("|", "\\|").replace("\n", " ").replace("\r", " ")


def _trim_trailing_blank(rows: list[list[str]]) -> list[list[str]]:
    while rows and not any(c.strip() for c in rows[-1]):
        rows.pop()
    return rows
