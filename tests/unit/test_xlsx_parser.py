"""xlsx parser unit tests."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from voitta_rag_enterprise.services.parsers.xlsx_parser import XlsxParser


def _build_workbook(tmp_path: Path) -> Path:
    wb = Workbook()
    s1 = wb.active
    s1.title = "Sales"
    s1.append(["Region", "Revenue"])
    s1.append(["EMEA", 100])
    s1.append(["APAC", 200])

    s2 = wb.create_sheet("Notes")
    s2.append(["Note"])
    s2.append(["Pipe | char survives"])

    hidden = wb.create_sheet("Hidden")
    hidden.append(["secret"])
    hidden.sheet_state = "hidden"

    out = tmp_path / "book.xlsx"
    wb.save(out)
    return out


def test_xlsx_parser_emits_one_section_per_visible_sheet(tmp_path: Path) -> None:
    p = _build_workbook(tmp_path)
    result = XlsxParser().parse(p)
    assert result.success
    md = result.content
    assert "## Sheet: Sales" in md
    assert "## Sheet: Notes" in md
    # Hidden sheet is skipped.
    assert "## Sheet: Hidden" not in md
    # Sales table renders as a pipe table with Region as header.
    assert "| Region | Revenue |" in md
    assert "| EMEA | 100 |" in md
    # Pipe in cell is escaped.
    assert "Pipe \\| char survives" in md


def test_xlsx_parser_handles_unparseable_file(tmp_path: Path) -> None:
    p = tmp_path / "bad.xlsx"
    p.write_bytes(b"not a real xlsx")
    result = XlsxParser().parse(p)
    assert not result.success
    assert "xlsx open failed" in (result.error or "")


def test_xlsx_parser_extensions() -> None:
    parser = XlsxParser()
    assert parser.can_parse(Path("a.xlsx"))
    assert parser.can_parse(Path("a.XLSM"))
    assert not parser.can_parse(Path("a.csv"))
